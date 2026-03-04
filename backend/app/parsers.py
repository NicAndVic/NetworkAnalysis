from __future__ import annotations

import configparser
import io
import logging
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from .models import CloudService, Device, ISP, Link, Topology
from .vendor_parsers import parse_aruba, parse_cisco, parse_procurve
from .vendor_parsers.common import ParsedDevice

logger = logging.getLogger(__name__)


def _normalize_speed(v: str) -> str:
    t = v.strip().upper().replace(" ", "")
    t = t.replace("GBPS", "G").replace("MBPS", "M")
    m = re.search(r"(100|40|25|10|1)(G|M)", t)
    return f"{m.group(1)}{m.group(2)}" if m else "1G"


def _detect_vendor(content: str, filename: str) -> str:
    c = content.lower()
    f = filename.lower()
    if "cisco ios" in c or "show cdp neighbors" in c or "_cdp_" in f:
        return "cisco"
    if "procurve" in c or "show lldp info remote-device detail" in c or "trk" in c:
        return "procurve"
    if "arubaos" in c or "show lldp neighbors detail" in c:

DEVICE_ALLOWED = {"switch", "firewall", "router", "server", "ap"}
IGNORE_TYPES = {"phone", "camera", "printer", "workstation", "endpoint"}


def _normalize_speed(v: str) -> str:
    t = v.strip().upper().replace("MBPS", "M").replace("GBPS", "G")
    t = t.replace(" ", "")
    if t in {"1000M", "1GB", "1GIG"}:
        return "1G"
    if t in {"100M", "10M", "1G", "10G", "25G", "40G", "100G"}:
        return t
    m = re.search(r"(\d+)(G|M)", t)
    return f"{m.group(1)}{m.group(2)}" if m else "1G"


def _detect_vendor(text: str) -> str:
    lowered = text.lower()
    if "cisco ios" in lowered or "show cdp neighbors" in lowered:
        return "cisco"
    if "procurve" in lowered or "show lldp info remote-device" in lowered:
        return "hp procurve"
    if "arubaos" in lowered or "show lldp neighbors detail" in lowered:
        return "aruba"
    return "unknown"


def _extract_hostname(content: str, filename: str) -> str:
    for rx in [r"^hostname\s+(\S+)", r"System\s+Name\s*:\s*(\S+)"]:
        m = re.search(rx, content, re.MULTILINE | re.IGNORECASE)
        if m:
            return m.group(1)
    # fallback: core1_show_xxx.txt -> core1
    stem = Path(filename).stem
    if "_show_" in stem:
        return stem.split("_show_", 1)[0]
    return stem


def _parse_by_vendor(vendor: str, content: str, fallback_hostname: str) -> ParsedDevice:
    if vendor == "cisco":
        return parse_cisco(content, fallback_hostname)
    if vendor == "procurve":
        return parse_procurve(content, fallback_hostname)
    if vendor == "aruba":
        return parse_aruba(content, fallback_hostname)
    # unknown: attempt cisco parser for broadest compatibility
    return parse_cisco(content, fallback_hostname)


def _merge_device_data(records: List[ParsedDevice]) -> ParsedDevice:
    base = records[0]
    for r in records[1:]:
        base.mgmt_ips = sorted(set(base.mgmt_ips + r.mgmt_ips))
        if not base.model:
            base.model = r.model
        base.roles = sorted(set(base.roles + r.roles))
        base.stp_root = base.stp_root or r.stp_root
        base.stp_blocked_ports = sorted(set(base.stp_blocked_ports + r.stp_blocked_ports))
        base.neighbors.extend(r.neighbors)
        base.vlans = sorted(set(base.vlans + r.vlans))
        base.routes = sorted(set(base.routes + r.routes))
        base.dhcp_scopes = sorted(set(base.dhcp_scopes + r.dhcp_scopes))
        if not base.stack_id and r.stack_id:
            base.stack_id = r.stack_id
    return base


def _add_or_update_device(topo: Topology, pd: ParsedDevice) -> None:
    d = Device(
        id=pd.hostname,
        hostname=pd.hostname,
        device_type="switch",
        vendor=pd.vendor,
        model=pd.model,
        mgmt_ips=pd.mgmt_ips,
        roles=pd.roles,
        stp_root=pd.stp_root,
        stack_id=pd.stack_id,
    )
    topo.devices[d.id] = d


def _dedupe_links(links: List[Link]) -> List[Link]:
    out = {}
    for l in links:
        a, b = sorted([l.src, l.dst])
        ap, bp = (l.src_port, l.dst_port) if a == l.src else (l.dst_port, l.src_port)
        key = (a, b, ap, bp, l.link_type, l.trunk_id, l.media, l.speed)
        out[key] = l
    return list(out.values())


def _enforce_ha_redundancy(topo: Topology, firewall_names: List[str], enabled: bool, disable_override: bool) -> None:
    if not enabled or len(firewall_names) < 2 or disable_override:
        return
    fw_set = set(firewall_names)
    all_links = topo.links[:]
    additions: List[Link] = []
    for link in all_links:
        if link.src in fw_set or link.dst in fw_set:
            core = link.dst if link.src in fw_set else link.src
            fw = link.src if link.src in fw_set else link.dst
            if core not in topo.devices or topo.devices[core].device_type != "switch":
                continue
            for target_fw in firewall_names:
                if target_fw == fw:
                    continue
                exists = any(
                    (l.src == core and l.dst == target_fw) or (l.src == target_fw and l.dst == core)
                    for l in topo.links + additions
                )
                if not exists:
                    additions.append(
                        Link(
                            src=core,
                            dst=target_fw,
                            src_port=link.src_port,
                            dst_port=link.dst_port,
                            speed=link.speed,
                            media=link.media,
                            link_type=link.link_type,
                        )
                    )
    topo.links.extend(additions)
def _parse_blocks(content: str) -> Dict[str, str]:
    # Expected raw bundle format: -- show command --\n<output>
    sections: Dict[str, List[str]] = defaultdict(list)
    current = "raw"
    for line in content.splitlines():
        if line.startswith("-- ") and line.endswith(" --"):
            current = line[3:-3].strip().lower()
            continue
        sections[current].append(line)
    return {k: "\n".join(v) for k, v in sections.items()}


def _parse_device(content: str, fallback_id: str) -> Tuple[Device, List[Link], List[str], List[str], List[str]]:
    blocks = _parse_blocks(content)
    raw = blocks.get("raw", content)
    vendor = _detect_vendor(content)

    hostname = fallback_id
    if m := re.search(r"hostname\s+(\S+)", raw, re.IGNORECASE):
        hostname = m.group(1)
    elif m := re.search(r"System\s+Name\s*:\s*(\S+)", content):
        hostname = m.group(1)

    model = ""
    if m := re.search(r"Model\s*(?:number)?\s*[:#]?\s*([\w\-]+)", content, re.IGNORECASE):
        model = m.group(1)

    mgmt_ips = sorted(set(re.findall(r"\b(?:10|172|192)\.\d+\.\d+\.\d+\b", content)))[:2]
    roles: List[str] = []
    device_type = "switch"
    stp_root = bool(re.search(r"This bridge is the root|Root this switch", content, re.IGNORECASE))
    stack_id = None
    if re.search(r"Switch\s+\d+\s+Provisioned|Stack member", content, re.IGNORECASE):
        stack_id = "core-stack" if hostname.startswith("CORE") else f"stack-{hostname}"

    if re.search(r"ip routing|routing enabled", raw, re.IGNORECASE):
        roles.append("L3")
    if re.search(r"dhcp pool|dhcp-server pool", raw, re.IGNORECASE):
        roles.append("DHCP")

    lldp = blocks.get("show lldp neighbors detail", "") + "\n" + blocks.get("show lldp info remote-device detail", "")
    cdp = blocks.get("show cdp neighbors detail", "")
    neighbors = lldp + "\n" + cdp

    links: List[Link] = []
    # Canonical neighbor lines used in samples but represent raw extraction product.
    for m in re.finditer(
        r"Local\s+Port\s*:\s*(\S+).*?Neighbor\s*:\s*(\S+).*?Neighbor\s+Port\s*:\s*(\S+).*?Type\s*:\s*([\w\-]+).*?Speed\s*:\s*([^\n]+).*?Media\s*:\s*([\w\-]+)(?:.*?Trunk\s*:\s*(\S+))?",
        neighbors,
        re.IGNORECASE | re.DOTALL,
    ):
        local_p, remote, remote_p, remote_type, speed, media, trunk = m.groups()
        rt = remote_type.lower()
        if rt in IGNORE_TYPES or rt not in DEVICE_ALLOWED:
            continue
        links.append(
            Link(
                src=hostname,
                dst=remote,
                src_port=local_p,
                dst_port=remote_p,
                speed=_normalize_speed(speed),
                media=media.lower(),
                link_type=(trunk or "normal").lower(),
                trunk_id=trunk,
            )
        )

    blocked_ports = set(re.findall(r"(?:BLK|Blocking)\s+(\S+)", blocks.get("show spanning-tree", ""), re.IGNORECASE))
    for link in links:
        if link.src_port in blocked_ports:
            link.stp_blocked = True

    vlan_lines = [f"{hostname}: VLAN {vid} {name}" for vid, name in re.findall(r"^\s*(\d+)\s+([\w\-]+)", blocks.get("show vlan", "") + "\n" + blocks.get("show vlan brief", ""), re.MULTILINE)]
    route_lines = [f"{hostname}: {p} {r}" for p, r in re.findall(r"^(S|O|C|L)\s+([^\n]+)", blocks.get("show ip route", ""), re.MULTILINE)]
    dhcp_lines = [f"{hostname}: {x.strip()}" for x in re.findall(r"Pool\s+([^\n]+)", raw + "\n" + blocks.get("show ip dhcp pool", ""), re.IGNORECASE)]

    d = Device(
        id=hostname,
        hostname=hostname,
        device_type=device_type,
        vendor=vendor,
        model=model,
        mgmt_ips=mgmt_ips,
        roles=sorted(set(roles)),
        stp_root=stp_root,
        stack_id=stack_id,
    )
    return d, links, vlan_lines, dhcp_lines, route_lines


def parse_ini(ini_bytes: bytes, topo: Topology) -> None:
    parser = configparser.ConfigParser()
    parser.read_string(ini_bytes.decode("utf-8"))

    fw_names: List[str] = []
    ha_enabled = False
    redundancy_disabled = False
    if parser.has_section("firewall"):
        ha_enabled = parser.getboolean("firewall", "ha", fallback=False)
        fw_names = [v.strip() for v in parser.get("firewall", "nodes", fallback="FW-A,FW-B").split(",") if v.strip()]
        redundancy_disabled = parser.getboolean("firewall", "disable_core_redundancy", fallback=False)
    if parser.has_section("firewall"):
        ha = parser.getboolean("firewall", "ha", fallback=False)
        fw_names = [v.strip() for v in parser.get("firewall", "nodes", fallback="FW-A,FW-B").split(",") if v.strip()]
        for idx, name in enumerate(fw_names):
            topo.add_device(
                Device(
                    id=name,
                    hostname=name,
                    device_type="firewall",
                    model=parser.get("firewall", "model", fallback="Firewall"),
                    ha_cluster="ha-fw" if ha_enabled else None,
                    ha_role=parser.get("firewall", f"role_{idx+1}", fallback=("active" if idx == 0 else "standby")),
                )
            )
        if ha_enabled and parser.getboolean("firewall", "sync_link", fallback=True) and len(fw_names) > 1:
                    ha_cluster="ha-fw" if ha else None,
                    ha_role=parser.get("firewall", f"role_{idx+1}", fallback=("active" if idx == 0 else "standby")),
                )
            )
        if ha and parser.getboolean("firewall", "sync_link", fallback=True) and len(fw_names) > 1:
            topo.add_link(Link(src=fw_names[0], dst=fw_names[1], speed="10G", media="stacking", link_type="ha-sync"))

    topo.add_device(Device(id="Internet", hostname="Internet", device_type="internet", model="Cloud"))

    for sec in parser.sections():
        if sec.startswith("isp:"):
            name = sec.split(":", 1)[1]
            isp = ISP(
                name=name,
                circuit_id=parser.get(sec, "circuit_id", fallback=""),
                media=parser.get(sec, "media", fallback="fiber").lower(),
                speed=_normalize_speed(parser.get(sec, "speed", fallback="1G")),
            )
            topo.isps.append(isp)
            topo.add_device(Device(id=f"ISP:{name}", hostname=name, device_type="isp", model=isp.circuit_id))
            topo.add_link(Link(src="Internet", dst=f"ISP:{name}", speed=isp.speed, media=isp.media, link_type="internet-uplink"))
            for fw in fw_names:
                topo.add_link(Link(src=f"ISP:{name}", dst=fw, speed=isp.speed, media=isp.media, link_type="wan"))

        if sec.startswith("cloud:"):
            name = sec.split(":", 1)[1]
            provider = parser.get(sec, "provider", fallback="Other")
            direct = parser.getboolean(sec, "direct_to_firewall", fallback=False)
            topo.cloud_services.append(CloudService(name=name, provider=provider, direct_to_firewall=direct))
            topo.add_device(Device(id=f"CLOUD:{name}", hostname=name, device_type="cloud", model=provider))
            if direct and fw_names:
                topo.add_link(Link(src=f"CLOUD:{name}", dst=fw_names[0], speed="1G", media="fiber", link_type="cloud-direct"))
            else:
                topo.add_link(Link(src=f"CLOUD:{name}", dst="Internet", speed="1G", media="fiber", link_type="cloud-via-internet"))

    _enforce_ha_redundancy(topo, fw_names, ha_enabled, redundancy_disabled)


def parse_excel(excel_bytes: bytes, topo: Topology) -> None:
    wb = load_workbook(io.BytesIO(excel_bytes))
    if "devices" in wb.sheetnames:
        ws = wb["devices"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            device = Device(
                id=str(row[0]),
                hostname=str(row[0]),
                device_type=str(row[1] or "switch").lower(),
                model=str(row[2] or ""),
                mgmt_ips=[str(row[3])] if row[3] else [],
            )
            topo.devices.setdefault(device.id, device)
    if "links" in wb.sheetnames:
        ws = wb["links"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            topo.add_link(Link(src=str(row[0]), dst=str(row[1]), speed=_normalize_speed(str(row[2] or "1G")), media=str(row[3] or "copper").lower(), link_type=str(row[4] or "normal").lower()))
            topo.add_link(
                Link(
                    src=str(row[0]),
                    dst=str(row[1]),
                    speed=_normalize_speed(str(row[2] or "1G")),
                    media=str(row[3] or "copper").lower(),
                    link_type=str(row[4] or "normal").lower(),
                )
            )


def parse_zip(zip_bytes: bytes) -> Topology:
    topo = Topology()

    by_host: Dict[str, List[ParsedDevice]] = defaultdict(list)
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for member in zf.namelist():
            if not member.lower().endswith(".txt"):
                continue
            content = zf.read(member).decode("utf-8", errors="ignore")
            host = _extract_hostname(content, member)
            vendor = _detect_vendor(content, member)
            try:
                parsed = _parse_by_vendor(vendor, content, host)
                by_host[parsed.hostname].append(parsed)
            except Exception as exc:
                logger.warning("could not classify/parse file %s: %s", member, exc)

    if not by_host:
        logger.warning("No parsable files found in ZIP")

    raw_links: List[Link] = []
    for _, records in by_host.items():
        pd = _merge_device_data(records)
        _add_or_update_device(topo, pd)
        for n in pd.neighbors:
            raw_links.append(
                Link(
                    src=pd.hostname,
                    dst=n.remote_host,
                    src_port=n.local_intf,
                    dst_port=n.remote_intf,
                    speed=n.speed,
                    media=n.media,
                    link_type=(n.trunk_id or "normal").lower(),
                    trunk_id=n.trunk_id,
                    stp_blocked=n.local_intf in pd.stp_blocked_ports,
                )
            )
        topo.vlan_lines.extend([f"{pd.hostname}: {v}" for v in pd.vlans])
        topo.dhcp_lines.extend([f"{pd.hostname}: {d}" for d in pd.dhcp_scopes])
        topo.route_lines.extend([f"{pd.hostname}: {r}" for r in pd.routes])
    raw_links: List[Link] = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for member in zf.namelist():
            if not member.endswith(".txt"):
                continue
            content = zf.read(member).decode("utf-8", errors="ignore")
            device, links, vlans, dhcp, routes = _parse_device(content, Path(member).stem)
            topo.add_device(device)
            raw_links.extend(links)
            topo.vlan_lines.extend(vlans)
            topo.dhcp_lines.extend(dhcp)
            topo.route_lines.extend(routes)

    grouped: Dict[Tuple[str, str, str], List[Link]] = defaultdict(list)
    for lk in raw_links:
        a, b = sorted([lk.src, lk.dst])
        key = lk.trunk_id or lk.link_type
        grouped[(a, b, (key or "normal").lower())].append(lk)

    final_links: List[Link] = []
    for (_, _, key), links in grouped.items():
        if (key.startswith(("po", "trk", "lag")) or key in {"port-channel", "lacp"}) and len(links) >= 2:
        grouped[(a, b, key)].append(lk)

    for (_, _, key), links in grouped.items():
        if (key.lower().startswith(("po", "trk", "lag")) or key.lower() in {"port-channel", "lacp"}) and len(links) >= 2:
            seed = links[0]
            seed.members = len(links)
            seed.link_type = "trunk"
            seed.trunk_id = key
            final_links.append(seed)
        else:
            final_links.extend(links)

    topo.links.extend(_dedupe_links(final_links))
            topo.add_link(seed)
        else:
            for lk in links:
                lk.link_type = "normal" if lk.link_type.lower().startswith(("po", "trk", "lag")) and len(links) < 2 else lk.link_type
                topo.add_link(lk)

    # dedupe vlan/dhcp/routes
    topo.vlan_lines = sorted(set(topo.vlan_lines))
    topo.dhcp_lines = sorted(set(topo.dhcp_lines))
    topo.route_lines = sorted(set(topo.route_lines))
    return topo
